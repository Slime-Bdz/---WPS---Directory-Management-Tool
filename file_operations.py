"""
file_operations.py

该模块包含所有核心的文件操作逻辑，包括多进程的文件查找、
多线程的文件复制以及生成 Excel 报告。
"""
import sys
import os
import shutil
import traceback
import re
from pathlib import Path
import concurrent.futures
from PyQt5.QtCore import QObject, pyqtSignal
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from fuzzywuzzy import fuzz
import pandas as pd

# ----------------------------------------------------------------------
# 路径管理 - 在打包后也能够正确找到资源文件
# ----------------------------------------------------------------------
def resource_path(relative_path):
    """
    获取资源文件的绝对路径，以兼容 PyInstaller 打包后的环境。
    """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# ----------------------------------------------------------------------
# 文件系统设置 - 自动创建资源文件夹和默认 Excel 文件
# ----------------------------------------------------------------------
def setup_excel_files():
    """
    检查并创建程序所需的 resources 文件夹和默认 Excel 文件。
    如果文件不存在，将创建包含默认表头的空文件。
    """
    resources_dir = resource_path('resources')
    file_list_path = os.path.join(resources_dir, 'file_list.xlsx')
    file_list_updated_path = os.path.join(resources_dir, 'file_list_updated.xlsx')

    # 1. 检查并创建 resources 文件夹
    if not os.path.exists(resources_dir):
        os.makedirs(resources_dir)
        print(f"Created resources directory at: {resources_dir}")

    # 2. 检查并创建 file_list.xlsx
    if not os.path.exists(file_list_path):
        print(f"File not found: {file_list_path}. Creating new file...")
        df = pd.DataFrame(columns=['文件名'])
        df.to_excel(file_list_path, index=False)
        print("Created default file_list.xlsx with a header.")

    # 3. 检查并创建 file_list_updated.xlsx
    if not os.path.exists(file_list_updated_path):
        print(f"File not found: {file_list_updated_path}. Creating new file...")
        df = pd.DataFrame(columns=['文件名'])
        df.to_excel(file_list_updated_path, index=False)
        print("Created default file_list_updated.xlsx with a header.")


def _scan_root_process(root_dir, names_to_find_set, match_mode, min_fuzzy_score):
    """
    一个独立的、可被多进程调用的函数，用于在单个根目录中查找文件。
    该函数只接受可被序列化的参数。
    """
    found_files = {}

    if not os.path.exists(root_dir):
        return found_files

    # 预处理，提高查找效率
    if match_mode == 'regex':
        try:
            # 预编译所有正则表达式
            regex_patterns = {name: re.compile(name) for name in names_to_find_set}
        except re.error:
            # 如果正则表达式无效，返回空结果
            return found_files

    for dirpath, _, filenames in os.walk(root_dir):
        current_dir_files_set = set(filenames)

        if match_mode == 'exact':
            for name_to_find in names_to_find_set:
                if name_to_find not in found_files:
                    for filename in current_dir_files_set:
                        if name_to_find in filename:
                            found_files[name_to_find] = str(Path(dirpath) / filename)
                            break
        elif match_mode == 'fuzzy':
            for filename in current_dir_files_set:
                for name_to_find in names_to_find_set:
                    if name_to_find not in found_files and fuzz.ratio(name_to_find, filename) >= min_fuzzy_score:
                        found_files[name_to_find] = str(Path(dirpath) / filename)
                        break
        elif match_mode == 'regex':
            for filename in current_dir_files_set:
                for name_to_find, pattern in regex_patterns.items():
                    if name_to_find not in found_files and pattern.search(filename):
                        found_files[name_to_find] = str(Path(dirpath) / filename)
                        break

        # 如果已经找到所有文件，则提前退出
        if len(found_files) == len(names_to_find_set):
            break

    return found_files


class SearchWorker(QObject):
    """
    一个在独立线程中执行搜索和复制任务的工作者类。
    """
    finished = pyqtSignal()
    success = pyqtSignal(str)
    failed = pyqtSignal(str)
    progress = pyqtSignal(int, int, str)

    def __init__(self, excel_path, target_dir, roots, updated_excel_path, match_mode='exact', min_fuzzy_score=85):
        """初始化工作者。"""
        super().__init__()
        self.excel_path = excel_path
        self.target_dir = target_dir
        self.roots = roots
        self.updated_excel_path = updated_excel_path
        self.match_mode = match_mode
        self.min_fuzzy_score = min_fuzzy_score
        self._is_stopped = False
        self._executor = None

    def stop(self):
        """停止当前任务。"""
        self._is_stopped = True
        self.failed.emit("用户请求取消任务，正在停止...")
        if self._executor:
            self._executor.shutdown(wait=False, cancel_futures=True)

    def run(self):
        """开始执行任务。"""
        try:
            self._work()
        except Exception:
            traceback.print_exc()
            self.failed.emit("任务执行出错，请检查日志。")
        finally:
            self.finished.emit()

    def _copy_single_file(self, name_to_find, src_path, target_dir):
        """将单个文件或目录复制到目标文件夹。"""
        if self._is_stopped:
            return {'status': 'stopped', 'message': "任务已中断。", 'name': name_to_find}

        if src_path and os.path.exists(src_path):
            dst_name = os.path.basename(src_path)
            dst = os.path.join(target_dir, dst_name)
            try:
                if os.path.isfile(src_path):
                    shutil.copy2(src_path, dst)
                else:
                    if os.path.exists(dst):
                        shutil.rmtree(dst)
                    shutil.copytree(src_path, dst)
                return {'status': 'success', 'message': f"✅ 已复制: {dst_name}", 'name': name_to_find}
            except Exception as e:
                return {'status': 'failed', 'message': f"❌ 复制失败 ({name_to_find}): {e}", 'name': name_to_find}
        else:
            return {'status': 'failed', 'message': f"❌ 未找到: {name_to_find}", 'name': name_to_find}

    def _find_files_in_roots(self, names_to_find_set):
        """并行扫描多个根目录并汇总结果。"""
        combined_found_files = {}
        total_roots = len(self.roots)
        completed_roots = 0

        with concurrent.futures.ProcessPoolExecutor(max_workers=os.cpu_count() or 4) as executor:
            self._executor = executor
            futures = {
                executor.submit(_scan_root_process, root, names_to_find_set, self.match_mode, self.min_fuzzy_score): root
                for root in self.roots
            }

            for future in concurrent.futures.as_completed(futures):
                if self._is_stopped:
                    executor.shutdown(wait=False, cancel_futures=True)
                    break

                try:
                    result = future.result()
                    for name, path in result.items():
                        if name not in combined_found_files:
                            combined_found_files[name] = path
                            self.success.emit(f"🔍 找到文件: {name}")
                except Exception as e:
                    root_dir = futures[future]
                    self.failed.emit(f"❌ 扫描目录 {root_dir} 发生错误: {e}")

                completed_roots += 1
                search_progress_value = int((completed_roots / total_roots) * 70)
                self.progress.emit(search_progress_value, 100, f"🔎 正在扫描: {completed_roots}/{total_roots} 个目录")

                if len(combined_found_files) == len(names_to_find_set):
                    self.success.emit("✅ 已找到所有文件，提前结束搜索。")
                    executor.shutdown(wait=False, cancel_futures=True)
                    break

        self._executor = None
        return combined_found_files

    def _work(self):
        """任务主流程：加载 Excel, 查找文件, 复制文件, 生成报告。"""
        self.progress.emit(0, 100, "⚙️ 正在初始化...")
        os.makedirs(self.target_dir, exist_ok=True)

        try:
            # 这里不需要修改，你的主程序会调用 setup_excel_files 来确保文件存在
            wb_original = load_workbook(self.excel_path)
            ws_original = wb_original.active

            names_to_find = [
                str(row[0]).strip()
                for row in ws_original.iter_rows(min_row=2, values_only=True)
                if row and row[0] is not None
            ]
            names_to_find = [name for name in names_to_find if name]
            names_to_find_set = set(names_to_find)

        except Exception as e:
            self.failed.emit(f"❌ 无法读取 Excel 文件: {self.excel_path} - {e}")
            return

        if not names_to_find_set:
            self.progress.emit(100, 100, "⚠️ Excel 中未找到文件名。")
            return

        self.success.emit(f"🔎 开始在 {len(self.roots)} 个目录中查找 {len(names_to_find)} 个文件...")
        found_files = self._find_files_in_roots(names_to_find_set)
        self.progress.emit(70, 100, "✅ 搜索阶段完成，准备复制文件...")

        if self._is_stopped:
            self.failed.emit("任务已中断。")
            return

        copy_results = self._copy_files(names_to_find, found_files)
        
        if not self._is_stopped:
            self._finalize_excel_report(self.updated_excel_path, names_to_find, copy_results)
            self.progress.emit(100, 100, "任务完成。")
            self.success.emit(f"✅ 已保存更新表：{Path(self.updated_excel_path).name}")
        else:
            self.failed.emit("任务已中断。")

    def _copy_files(self, names_to_find, found_files):
        """使用多线程复制文件。"""
        total_files_to_process = len(names_to_find)
        if total_files_to_process == 0:
            self.success.emit("没有需要复制的文件。")
            return []

        copied_count = 0
        copy_results = []
        self.progress.emit(70, 100, "📁 正在并发复制文件...")

        with concurrent.futures.ThreadPoolExecutor(max_workers=os.cpu_count() * 2 or 4) as executor:
            futures = [executor.submit(self._copy_single_file, name, found_files.get(name), self.target_dir)
                       for name in names_to_find]
            
            for future in concurrent.futures.as_completed(futures):
                if self._is_stopped:
                    executor.shutdown(wait=False, cancel_futures=True)
                    break
                
                copied_count += 1
                try:
                    result = future.result()
                    copy_results.append(result)
                    if result['status'] == 'success':
                        self.success.emit(result['message'])
                    elif result['status'] == 'failed':
                        self.failed.emit(result['message'])
                except Exception as e:
                    self.failed.emit(f"❌ 任务处理异常: {e}")

                copy_progress_value = 70 + int((copied_count / total_files_to_process) * 30)
                self.progress.emit(copy_progress_value, 100, f"🚀 正在复制文件: {copied_count}/{total_files_to_process}")
        
        return copy_results

    def _finalize_excel_report(self, updated_excel_path, names_to_find, copy_results):
        """生成并保存最终的 Excel 报告。"""
        try:
            if not os.path.exists(updated_excel_path):
                wb = Workbook()
                ws = wb.active
                ws.cell(row=1, column=1, value="文件名")
                ws.cell(row=1, column=2, value="状态")
                wb.save(updated_excel_path)
            
            wb = load_workbook(updated_excel_path)
            ws = wb.active

            results_map = {res['name']: res for res in copy_results}

            for _ in range(len(names_to_find) - ws.max_row):
                ws.append(['', ''])

            for idx, name_to_find in enumerate(names_to_find):
                row_index = idx + 2
                result = results_map.get(name_to_find)

                fill = None
                status_text = ""
                if result:
                    status = result['status']
                    if status == 'success':
                        fill = PatternFill(fill_type='solid', start_color='00FF00', end_color='00FF00')
                        status_text = "✅ 已找到"
                    elif status == 'failed':
                        fill = PatternFill(fill_type='solid', start_color='FFC0CB', end_color='FFC0CB')
                        status_text = "❌ 未找到或复制失败"
                else:
                    fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
                    status_text = "❌ 未找到"

                ws.cell(row=row_index, column=1, value=name_to_find)
                cell_status = ws.cell(row=row_index, column=2, value=status_text)
                if fill:
                    cell_status.fill = fill
            
            wb.save(updated_excel_path)
        except Exception as e:
            self.failed.emit(f"❌ 无法保存更新的 Excel 报告: {e}")
            traceback.print_exc()