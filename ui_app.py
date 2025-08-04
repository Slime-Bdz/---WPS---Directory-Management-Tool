import os
import sys
import shutil
import traceback
import tempfile
from pathlib import Path
from PyQt5.QtCore import QThread, pyqtSignal, QObject, Qt, QAbstractTableModel, QModelIndex, QUrl
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QFileDialog,
    QTextEdit, QLabel, QSplitter, QGroupBox, QLineEdit, QTabWidget, QTableView,
    QHeaderView, QAbstractItemView
)
from PyQt5.QtGui import QDesktopServices, QFont, QClipboard, QKeySequence
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# -------------------------------------------------
# 自动释放资源
# -------------------------------------------------
def resource_path(rel):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, rel)
    return os.path.join(os.path.abspath('.'), rel)

def ensure_embedded_excels():
    # 将文件创建在当前工作目录
    current_dir = Path('.')
    
    excels = ['file_list.xlsx', 'file_list_updated.xlsx']
    paths = {}
    for name in excels:
        dst = current_dir / name
        if not dst.exists():
            # 如果文件在当前目录不存在，则从资源路径复制过来
            src = resource_path(name)
            shutil.copy2(src, dst)
        paths[name] = str(dst)
    return paths['file_list.xlsx'], paths['file_list_updated.xlsx']

# -------------------------------------------------
# Excel 表格模型（用于 QTableView）
# -------------------------------------------------
class ExcelTableModel(QAbstractTableModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.headers = []
        self.data_list = []

    def load(self, path):
        self.beginResetModel()
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.append(["文件名"])
            wb.save(path)
        df = pd.read_excel(path, dtype=str).fillna("")
        self.headers = df.columns.tolist()
        self.data_list = df.values.tolist()
        self.endResetModel()

    def save(self, path):
        df = pd.DataFrame(self.data_list, columns=self.headers or ["文件名"])
        df.to_excel(path, index=False)

    def rowCount(self, parent=QModelIndex()):
        return len(self.data_list)

    def columnCount(self, parent=QModelIndex()):
        # 确保即使没有数据，也至少返回 1 列（例如，如果只有 headers 被初始化）
        return len(self.headers) if self.headers else 1

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        if role == Qt.DisplayRole or role == Qt.EditRole:
            # 确保索引在有效范围内，并且行存在，列在当前行的实际列数范围内
            if index.row() < len(self.data_list) and \
               index.column() < len(self.data_list[index.row()]):
                return str(self.data_list[index.row()][index.column()])
        return None

    # 新增方法：确保模型中有足够的行
    def _ensure_row_exists(self, target_row_index):
        # 仅当目标行索引超出当前数据列表的末尾时才插入新行
        if target_row_index >= len(self.data_list):
            self.beginInsertRows(QModelIndex(), len(self.data_list), target_row_index)
            # 循环添加新行，直到达到目标行索引
            for _ in range(len(self.data_list), target_row_index + 1):
                # 新行应包含与当前列数相同的空字符串
                self.data_list.append([''] * self.columnCount())
            self.endInsertRows()

    def setData(self, index, value, role=Qt.EditRole):
        if not index.isValid():
            return False
        
        if role == Qt.EditRole:
            try:
                # 1. 确保目标行存在
                while index.row() >= len(self.data_list):
                    self.beginInsertRows(QModelIndex(), len(self.data_list), len(self.data_list))
                    initial_cols_for_new_row = self.columnCount()
                    self.data_list.append([''] * initial_cols_for_new_row)
                    self.endInsertRows()

                # 2. 确保表头有足够的列，并为所有现有行扩展这些列
                while index.column() >= len(self.headers):
                    new_col_name = f"Column{len(self.headers) + 1}"
                    self.headers.append(new_col_name)
                    # 遍历所有现有行，为它们添加新的空列
                    for r_idx in range(len(self.data_list)):
                        # 只对需要扩展的行进行操作，避免重复
                        if r_idx < len(self.data_list) and len(self.data_list[r_idx]) < len(self.headers):
                            self.data_list[r_idx].append('')
                    self.headerDataChanged.emit(Qt.Horizontal, len(self.headers) - 1, len(self.headers) - 1)
                
                # 3. 确保目标行（data_list[index.row()]）的内部列表有足够的列
                while index.row() < len(self.data_list) and index.column() >= len(self.data_list[index.row()]):
                    self.data_list[index.row()].append('')

                # 现在，可以安全地设置数据
                self.data_list[index.row()][index.column()] = str(value)
                self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
                return True
            except Exception as e:
                traceback.print_exc()
                return False
        return False

# -------------------------------------------------
# 自定义 QTableView
# -------------------------------------------------
class CustomTableView(QTableView):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.setSelectionBehavior(QAbstractItemView.SelectItems)

    def keyPressEvent(self, event):
        if event.matches(QKeySequence.Copy):
            self.copySelected()
        elif event.matches(QKeySequence.Paste):
            self.pasteFromClipboard()
        elif event.key() == Qt.Key_Delete or event.key() == Qt.Key_Backspace:
            self.deleteSelected()
        else:
            super().keyPressEvent(event)

    def copySelected(self):
        selection = self.selectionModel().selectedIndexes()
        if not selection:
            return

        min_row = min(index.row() for index in selection)
        max_row = max(index.row() for index in selection)
        min_col = min(index.column() for index in selection)
        max_col = max(index.column() for index in selection)

        table_data = []
        for r in range(min_row, max_row + 1):
            row_data = []
            for c in range(min_col, max_col + 1):
                index = self.model().index(r, c)
                # 检查索引是否在模型有效范围内，并且只复制选中项
                if r < self.model().rowCount() and c < self.model().columnCount():
                     if index in selection:
                        row_data.append(str(self.model().data(index, Qt.DisplayRole)))
                     else:
                        row_data.append('') # 未选中但仍在框选区域内的单元格，复制为空
                else:
                    row_data.append('') # 超出模型范围的单元格，复制为空
            table_data.append(row_data)

        text_to_copy = '\n'.join(['\t'.join(row) for row in table_data])
        QApplication.clipboard().setText(text_to_copy)

    def deleteSelected(self):
        selection = self.selectionModel().selectedIndexes()
        if not selection:
            return

        # 获取选区边界
        min_row = min(index.row() for index in selection)
        max_row = max(index.row() for index in selection)
        min_col = min(index.column() for index in selection)
        max_col = max(index.column() for index in selection)
        
        model = self.model()
        # 清除选中区域内的所有内容
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                index = model.index(r, c)
                if index.isValid():
                    model.setData(index, '', Qt.EditRole)

    def pasteFromClipboard(self):
        clipboard = QApplication.clipboard()
        text = clipboard.text()

        if not text:
            return
        
        if text.endswith('\n'):
            text = text[:-1]

        rows_data = [line.split('\t') for line in text.split('\n')]
        pasted_rows_count = len(rows_data)

        model = self.model() 
        selection_model = self.selectionModel()
        selected_indexes = selection_model.selectedIndexes()

        start_row = 0
        start_col = 0
        selected_rows_count = 0 
        clear_min_col = 0 
        clear_max_col = 0 

        if not selected_indexes:
            current_index = self.currentIndex()
            if current_index.isValid():
                start_row = current_index.row()
                start_col = current_index.column()
        else:
            min_row = min(index.row() for index in selected_indexes)
            max_row = max(index.row() for index in selected_indexes)
            min_col = min(index.column() for index in selected_indexes)
            max_col = max(index.column() for index in selected_indexes)

            start_row = min_row
            start_col = min_col
            selected_rows_count = max_row - min_row + 1
            
            clear_min_col = min_col 
            clear_max_col = max_col 

        model._ensure_row_exists(start_row + pasted_rows_count - 1) 

        for r_offset, row_values in enumerate(rows_data):
            target_row = start_row + r_offset
            for c_offset, value in enumerate(row_values):
                target_col = start_col + c_offset
                target_index = model.index(target_row, target_col)
                if target_index.isValid(): 
                    model.setData(target_index, value, Qt.EditRole)

        if selected_rows_count > 0 and pasted_rows_count < selected_rows_count:
            rows_to_clear_start = start_row + pasted_rows_count
            rows_to_clear_end = start_row + selected_rows_count - 1 

            actual_clear_end = min(rows_to_clear_end, model.rowCount() - 1)

            for r in range(rows_to_clear_start, actual_clear_end + 1):
                for c in range(clear_min_col, clear_max_col + 1):
                    index_to_clear = model.index(r, c)
                    if index_to_clear.isValid():
                        model.setData(index_to_clear, '', Qt.EditRole)


# -------------------------------------------------
# 业务线程（已加自动创建目标目录）
# -------------------------------------------------
class SearchWorker(QObject):
    finished = pyqtSignal()
    success  = pyqtSignal(str)
    failed   = pyqtSignal(str)

    def __init__(self, excel, target, roots):
        super().__init__()
        self.excel = excel
        self.target = target
        self.roots = roots

    def run(self):
        try:
            self._work()
        except Exception:
            traceback.print_exc()
            self.failed.emit("任务执行出错，请检查日志。")
        self.finished.emit()

    def _work(self):
        os.makedirs(self.target, exist_ok=True)
        df = pd.read_excel(self.excel)
        names = [str(x).strip() for x in df.iloc[:, 0]]
        found = {}
        for root in self.roots:
            if not os.path.exists(root):
                self.failed.emit(f"查找根目录不存在: {root}")
                continue
            for dp, _, fs in os.walk(root):
                for n in names:
                    if n in found:
                        continue
                    if n in fs or n in os.listdir(dp):
                        found[n] = os.path.join(dp, n)
                    else:
                        for item in fs + os.listdir(dp):
                            stem, ext = os.path.splitext(item)
                            if stem == n and ext:
                                found[n] = os.path.join(dp, item)
                                break

        wb = load_workbook(self.excel)
        ws = wb.active
        for idx, name in enumerate(names, 2):
            src = found.get(name)
            if src and os.path.exists(src):
                dst_name = os.path.basename(src) # 获取文件名
                dst = os.path.join(self.target, dst_name)
                try:
                    if os.path.isfile(src):
                        shutil.copy2(src, dst)
                    else:
                        if os.path.exists(dst):
                            shutil.rmtree(dst)
                        shutil.copytree(src, dst)
                    # 精简成功日志输出，只显示文件名
                    self.success.emit(f"✅ 已复制: {dst_name}")
                except Exception as e:
                    self.failed.emit(f"❌ 复制失败 ({name}): {e}")
                    ws.cell(row=idx, column=1).fill = PatternFill(fill_type='solid',
                                                                  start_color='FFC0CB',
                                                                  end_color='FFC0CB')
            else:
                self.failed.emit(f"❌ 未找到: {name}")
                ws.cell(row=idx, column=1).fill = PatternFill(fill_type='solid',
                                                              start_color='FFFF00',
                                                              end_color='FFFF00')

        updated_path = Path(self.excel).with_name('file_list_updated.xlsx')
        wb.save(updated_path)
        self.success.emit(f'✅ 已保存更新表：{updated_path.name}')

# -------------------------------------------------
# 主界面
# -------------------------------------------------
class UniApp(QWidget):
    SETTINGS_FILE = "last_paths.ini"

    def __init__(self):
        super().__init__()
        self.setWindowTitle('目录管理终端')
        self.resize(1000, 800)
        self.setWindowFlag(Qt.FramelessWindowHint)

        self.excel_le   = QLineEdit(self)
        self.target_le  = QLineEdit(self)
        self.root_le    = QLineEdit(self)

        self.excel_btn  = QPushButton('浏览...', self)
        self.target_btn = QPushButton('浏览...', self)
        self.root_btn   = QPushButton('浏览...', self)
        self.start_btn  = QPushButton('开始执行', self)
        # 新增按钮
        self.create_refresh_excels_btn = QPushButton('创建/刷新 Excel 表', self)

        self.model_origin  = ExcelTableModel()
        self.model_updated = ExcelTableModel()
        self.view_origin = CustomTableView(self)
        self.view_updated = CustomTableView(self)

        self.success_edit = QTextEdit(self)
        self.fail_edit    = QTextEdit(self)

        self.tab = QTabWidget()
        self.tab.addTab(self._build_work_tab(), "工作终端")
        self.tab.addTab(self._build_excel_tab(self.model_origin, "file_list.xlsx"), "file_list.xlsx")
        self.tab.addTab(self._build_excel_tab(self.model_updated, "file_list_updated.xlsx"), "file_list_updated.xlsx")
        self.tab.addTab(self._build_about_tab(), "关于")

        self._build_ui()
        self._signals()
        self._apply_styles()
        self.load_excels()
        self.load_settings()

    def _build_work_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)

        form = QGroupBox('路径设置')
        form_layout = QVBoxLayout(form)
        for title, le, btn in [('Excel 列表', self.excel_le, self.excel_btn),
                               ('目标文件夹', self.target_le, self.target_btn),
                               ('查找根目录', self.root_le, self.root_btn)]:
            h_layout = QHBoxLayout()
            h_layout.addWidget(QLabel(title))
            h_layout.addWidget(le)
            h_layout.addWidget(btn)
            form_layout.addLayout(h_layout)
        layout.addWidget(form)

        # 在布局中添加新按钮
        layout.addWidget(self.create_refresh_excels_btn)
        layout.addWidget(self.start_btn)
        return widget

    def _build_excel_tab(self, model, title):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        view = CustomTableView(self)
        view.setModel(model)
        view.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(view)
        save_btn = QPushButton(f"保存 {title}")
        save_btn.clicked.connect(lambda: self.save_excel(model, title))
        layout.addWidget(save_btn)
        return widget

    def _build_about_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        about_text = QTextEdit()
        about_text.setReadOnly(True)
        # 替换为新的简介文本
        about_text.setText(
            """
            <h2 style='color:#00FFFF;'>目录管理终端</h2>
            <p style='color:#E0E0E0;'>本工具旨在简化批量文件查找与整理流程，助您高效管理大量文件。</p>
            
            <h3 style='color:#00FFFF;'>使用流程：</h3>
            <ol>
                <li style='color:#E0E0E0;'><b>导入列表：</b> 导入包含目标文件名的 Excel 列表。</li>
                <li style='color:#E0E0E0;'><b>设定路径：</b> 设置文件查找的“根目录”和文件复制的“目标文件夹”。</li>
                <li style='color:#E0E0E0;'><b>一键执行：</b> 程序将自动搜索并复制文件，同时生成详细的执行报告。</li>
            </ol>
            
            <h3 style='color:#00FFFF;'>特色功能：</h3>
            <ul>
                <li style='color:#E0E0E0;'><b>Excel 驱动：</b> 通过 Excel 列表进行批量查找与复制，告别手动操作。</li>
                <li style='color:#E0E0E0;'><b>智能匹配：</b> 支持文件名及“词干”匹配，提高查找成功率。</li>
                <li style='color:#E0E0E0;'><b>实时报告：</b> 即时查看成功/失败日志，任务完成后自动生成带标记的更新版 Excel 报告。</li>
                <li style='color:#E0E0E0;'><b>内置编辑：</b> 直接在界面中编辑 Excel 列表，支持复制、粘贴、删除单元格内容。</li>
                <li style='color:#E0E0E0;'><b>提示：</b> 每次输入新表必须点击保存才能被应用</li>
                <li style='color:#E0E0E0;'><b>提示：</b> 首次使用最好点击创建/刷新Excel表</li>
            </ul>
            
            <p style='color:#E0E0E0;'>告别繁琐，让文件管理轻松高效！</p>
            """
        )
        layout.addWidget(about_text)
        
        # 调整 QTextEdit 的样式
        about_text.setStyleSheet("""
            QTextEdit {
                background: #1A1F36;
                border: 1px solid #00FFFF;
                border-radius: 6px;
                color: #E0E0E0; /* 文本颜色 */
                padding: 15px; /* 增加内边距 */
                font-family: "Segoe UI", "Microsoft YaHei", "Consolas";
                font-size: 10pt;
            }
        """)
        
        return widget

    def _build_ui(self):
        main = QVBoxLayout(self)

        title_bar = QHBoxLayout()
        title_label = QLabel("目录管理终端")
        title_label.setAlignment(Qt.AlignCenter)
        title_bar.addWidget(title_label)
        title_bar.addStretch()
        close_btn = QPushButton("X")
        close_btn.clicked.connect(self.close)
        title_bar.addWidget(close_btn)
        main.addLayout(title_bar)

        main.addWidget(self.tab)

        log_splitter = QSplitter(Qt.Horizontal)
        log_splitter.addWidget(self._log_group("成功日志", self.success_edit))
        log_splitter.addWidget(self._log_group("失败日志", self.fail_edit))
        main.addWidget(log_splitter)

    def _log_group(self, title, text_edit):
        group = QGroupBox(title)
        layout = QVBoxLayout(group)
        layout.addWidget(text_edit)
        return group

    def _signals(self):
        self.excel_btn.clicked.connect(lambda: self.choose_file(self.excel_le, 'Excel 文件 (*.xlsx)'))
        self.target_btn.clicked.connect(lambda: self.choose_folder(self.target_le))
        self.root_btn.clicked.connect(lambda: self.choose_folder(self.root_le))
        self.start_btn.clicked.connect(self.start_task)
        # 连接新按钮的信号
        self.create_refresh_excels_btn.clicked.connect(self._create_and_refresh_excels)

    def choose_file(self, line_edit, filt):
        path, _ = QFileDialog.getOpenFileName(self, '选择文件', filter=filt)
        if path:
            line_edit.setText(path)
            self.load_excels()

    def choose_folder(self, line_edit):
        path = QFileDialog.getExistingDirectory(self, '选择文件夹')
        if path:
            line_edit.setText(path)

    def open_excel(self, updated=False):
        excel_path, updated_path = ensure_embedded_excels()
        QDesktopServices.openUrl(QUrl.fromLocalFile(updated_path if updated else excel_path))

    def load_excels(self):
        excel_path, updated_path = ensure_embedded_excels()
        self.model_origin.load(excel_path)
        self.model_updated.load(updated_path)

    def save_excel(self, model, title):
        excel_path, updated_path = ensure_embedded_excels()
        path = excel_path if title == "file_list.xlsx" else updated_path
        model.save(path)
        self.success_edit.append(f"✅ {title} 已保存")

    # 新增方法：创建/刷新 Excel 表格并刷新内容
    def _create_and_refresh_excels(self):
        try:
            excel_path, updated_path = ensure_embedded_excels() # 确保文件存在或被创建
            self.load_excels() # 重新加载模型以刷新表格内容
            self.success_edit.append("✅ Excel 表格已检测并刷新。")
            
            # 自动设置 Excel 路径到输入框
            self.excel_le.setText(excel_path)

        except Exception as e:
            traceback.print_exc()
            self.fail_edit.append(f"❌ 创建/刷新 Excel 表格失败: {e}")

    def start_task(self):
        excel = self.excel_le.text()
        target = self.target_le.text()
        root = self.root_le.text()
        if not all([excel, target, root]):
            self.fail_edit.append("请确保Excel列表、目标文件夹和查找根目录都已设置！")
            return
        self.start_btn.setEnabled(False)
        self.success_edit.clear()
        self.fail_edit.clear()
        self.save_settings()

        self.thread = QThread()
        self.worker = SearchWorker(excel, target, [root])
        self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.run)
        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(lambda: self.start_btn.setEnabled(True))
        # 任务完成后，重新加载 Excel 模型，以更新界面
        self.worker.finished.connect(self.load_excels)
        self.worker.success.connect(lambda s: self.success_edit.append(s))
        self.worker.failed.connect(lambda s: self.fail_edit.append(s))
        self.thread.start()

    def save_settings(self):
        with open(self.SETTINGS_FILE, "w", encoding="utf-8") as f:
            f.write("\n".join([self.excel_le.text(), self.target_le.text(), self.root_le.text()]))

    def load_settings(self):
        # 始终确保 excel_le 路径指向当前目录的 file_list.xlsx
        excel_path, _ = ensure_embedded_excels()
        self.excel_le.setText(excel_path)

        try:
            with open(self.SETTINGS_FILE, "r", encoding="utf-8") as f:
                lines = f.read().splitlines()
                # 如果设置文件存在，则加载目标和根目录路径
                if len(lines) >= 3:
                    # lines[0] 是旧的 excel_le 路径，我们在这里已经用 current_dir 的路径覆盖了它
                    self.target_le.setText(lines[1])
                    self.root_le.setText(lines[2])
        except FileNotFoundError:
            # 如果设置文件不存在，则 excel_le 已经通过 ensure_embedded_excels() 设置正确
            pass

    def _apply_styles(self):
        self.setStyleSheet("""
/* 全局设置 - 深邃星空主题 */
QWidget {
    background: #0A0A1A; /* 深星空蓝背景 */
    color: #E0E0E0; /* 柔和的白色文本 */
    font-family: "Segoe UI", "Microsoft YaHei", "Consolas";
    font-size: 9pt;
}

/* 标题 - 高亮青色 */
QLabel {
    color: #00FFFF; /* 赛博朋克青色 */
    font-size: 14pt;
    font-weight: bold;
}

/* 按钮 - 霓虹渐变效果 */
QPushButton {
    /* 修正后的渐变：浅蓝色到深蓝色 */
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop: 0 #00c6ff, stop: 1 #0078ff);
    border: none;
    border-radius: 6px;
    padding: 6px 18px;
    color: #fff; /* 白色文字以保证对比度 */
    font-weight: bold;
}

QPushButton:hover {
    /* 修正后的悬停渐变：深蓝色到浅蓝色 */
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop: 0 #0078ff, stop: 1 #00c6ff);
}

/* 分组框 - 科技感线框 */
QGroupBox {
    border: 1px solid #00FFFF;
    border-radius: 8px;
    margin-top: 6px;
    font-weight: bold;
    color: #00FFFF; /* 标题文字同为青色 */
}

QGroupBox::title {
    subcontrol-origin: margin;
    left: 8px;
    padding: 0 3px 0 3px;
}

/* 输入框 - 终端风格 */
QLineEdit {
    border: 1px solid #00FFFF;
    border-radius: 4px;
    padding: 4px;
    background: #1A1F36; /* 较浅的深蓝背景 */
    color: #00FFFF; /* 青色输入文字 */
}

/* 日志输出框 - 经典绿色荧光字 */
QTextEdit {
    background: #1A1F36;
    border: 1px solid #00FFFF;
    border-radius: 6px;
    color: #39FF14; /* 终端荧光绿 */
    font-family: "Consolas", "Courier New";
}

/* 表格视图 - 数据矩阵风格 */
QTableView {
    background: #1A1F36;
    border: 1px solid #00FFFF;
    border-radius: 6px;
    color: #E0E0E0;
    gridline-color: #007F7F; /* 较暗的青色网格线 */
}

QTableView::item {
    padding: 4px 0;
    border: none;
}

QTableView::item:selected {
    background: #00FFFF; /* 选中时背景为高亮青色 */
    color: #0A0A1A; /* 文字变为深色以保证可读性 */
}

QTableView QTableCornerButton::section {
    background-color: #101428; /* 与表头背景色一致 */
    border: none;
}

/* 表头 - 控制台风格 */
QHeaderView {
    background: #101428; /* 更深的蓝色 */
    color: #E0E0E0;
    border: 1px solid #007F7F;
    border-radius: 4px;
    padding: 4px;
}

QHeaderView::section {
    background: #1A1F36;
    color: #00FFFF;
    border: none;
    padding: 4px;
}

/* 选项卡 - 全息界面风格 */
QTabWidget::pane {
    border: 1px solid #00FFFF; /* 添加边框以包裹内容 */
    border-top-left-radius: 0px;
    border-top-right-radius: 0px;
    border-bottom-left-radius: 8px;
    border-bottom-right-radius: 8px;
    margin-top: -1px; /* 负边距，使内容与tab bar的底部边框对齐 */
    background: #1A1F36; /* 确保内容区域有背景色 */
}

QTabBar::tab {
    background: transparent; /* 透明背景 */
    color: #00FFFF;
    padding: 8px 15px; /* 减小左右填充，使页签变窄 */
    margin: 0;
    border: 1px solid #007F7F; /* 暗青色边框 */
    border-bottom: none;
    border-top-left-radius: 4px;
    border-top-right-radius: 4px;
    min-width: 100px; /* 减小最小宽度，使页签变窄 */
    font-size: 10pt; /* 可以适当减小字体大小 */
    font-weight: bold;
}

QTabBar::tab:selected {
    background: #1A1F36; /* 选中时有背景色 */
    color: #00FFFF;
    border: 1px solid #00FFFF; /* 边框变为高亮青色 */
    border-bottom: 1px solid #1A1F36; /* 底部边框与内容区域融合 */
}

QTabBar::tab:!selected:hover {
    color: #FF00FF; /* 未选中项悬停时变为品红色 */
    border-color: #FF00FF;
}
""")

# -------------------------------------------------
if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = UniApp()
    win.show()
    sys.exit(app.exec_())